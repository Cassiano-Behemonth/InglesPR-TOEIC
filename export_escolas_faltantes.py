#!/usr/bin/env python3
import os
import json
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_xlsx():
    html_path = 'comparativo_toeic.html'
    if not os.path.exists(html_path):
        print(f"Erro: Arquivo '{html_path}' não encontrado.")
        return

    with open(html_path, 'r', encoding='utf-8') as f:
        content = f.read()

    match = re.search(r'const listaEscolas = (\[.*?\]);', content, re.DOTALL)
    if not match:
        print("Erro: 'listaEscolas' não encontrado em comparativo_toeic.html")
        return

    data = json.loads(match.group(1))
    df = pd.DataFrame(data)

    # Filtrar apenas as escolas que faltam cadastrar (status == 'confirmed_only')
    df_faltam = df[df['status'] == 'confirmed_only'].copy()
    
    # Ordenar por NRE, Cidade e Escola
    df_faltam.sort_values(by=['nre', 'cidade_planilha', 'escola'], inplace=True)
    
    # Criar pasta de trabalho do Excel
    wb = openpyxl.Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Navy
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    
    title_font = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
    subtitle_font = Font(name="Segoe UI", size=10, italic=True, color="475569")
    
    regular_font = Font(name="Segoe UI", size=10, color="1E293B")
    bold_font = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
    
    red_badge_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_badge_font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
    
    green_badge_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    green_badge_font = Font(name="Segoe UI", size=10, bold=True, color="166534")

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )

    # ----------------------------------------------------
    # TAB 1: Escolas Faltantes
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Escolas Faltantes"
    ws1.views.sheetView[0].showGridLines = True
    
    # Título
    ws1.cell(row=1, column=1, value="Escolas Confirmadas Pendentes de Cadastro no Formulário (TOEIC)").font = title_font
    ws1.cell(row=2, column=1, value=f"Total de escolas pendentes: {len(df_faltam)} escolas").font = subtitle_font
    
    headers1 = ["Item", "NRE", "Município", "Nome da Escola (Ofício)", "Estudantes Previstos", "Status no Formulário"]
    
    start_row = 4
    for col_num, header_title in enumerate(headers1, 1):
        cell = ws1.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    ws1.row_dimensions[start_row].height = 28

    current_row = start_row + 1
    for idx, (_, r) in enumerate(df_faltam.iterrows(), 1):
        c1 = ws1.cell(row=current_row, column=1, value=idx)
        c2 = ws1.cell(row=current_row, column=2, value=r['nre'])
        c3 = ws1.cell(row=current_row, column=3, value=r['cidade_planilha'])
        c4 = ws1.cell(row=current_row, column=4, value=r['escola'])
        c5 = ws1.cell(row=current_row, column=5, value=r['alunos'])
        c6 = ws1.cell(row=current_row, column=6, value="Pendente")

        c1.alignment = Alignment(horizontal="center")
        c2.alignment = Alignment(horizontal="left")
        c3.alignment = Alignment(horizontal="left")
        c4.alignment = Alignment(horizontal="left")
        c5.alignment = Alignment(horizontal="right")
        c6.alignment = Alignment(horizontal="center")
        c6.fill = red_badge_fill
        c6.font = red_badge_font

        for cell in [c1, c2, c3, c4, c5]:
            cell.font = regular_font
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = zebra_fill
        c6.border = thin_border
        
        ws1.row_dimensions[current_row].height = 20
        current_row += 1

    # Linha de Total
    ws1.cell(row=current_row, column=1, value="TOTAL").font = bold_font
    ws1.cell(row=current_row, column=1).alignment = Alignment(horizontal="center")
    ws1.cell(row=current_row, column=4, value=f"{len(df_faltam)} escolas pendentes").font = bold_font
    ws1.cell(row=current_row, column=5, value=f"=SUM(E{start_row+1}:E{current_row-1})").font = bold_font
    
    for c in range(1, 7):
        cell = ws1.cell(row=current_row, column=c)
        cell.fill = total_fill
        cell.border = thin_border

    # Auto-adjust column widths
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)
    ws1.column_dimensions['D'].width = 55 # Escola width

    # Autofilter
    ws1.auto_filter.ref = f"A{start_row}:F{current_row-1}"

    # ----------------------------------------------------
    # TAB 2: Resumo por NRE
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Resumo por NRE")
    ws2.views.sheetView[0].showGridLines = True
    
    ws2.cell(row=1, column=1, value="Resumo da Situação de Cadastro por NRE").font = title_font
    ws2.cell(row=2, column=1, value="Comparativo entre escolas confirmadas vs. escolas pendentes").font = subtitle_font

    headers2 = ["NRE", "Total Confirmadas", "Cadastradas (OK)", "Faltantes (Pendentes)", "% Pendente"]
    
    start_row = 4
    for col_num, header_title in enumerate(headers2, 1):
        cell = ws2.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[start_row].height = 28

    # Agrupar dados por NRE
    nre_summary = []
    all_nres = sorted(df['nre'].unique())
    for nre_val in all_nres:
        if not nre_val or nre_val == 'NRE indefinido':
            continue
        df_nre = df[df['nre'] == nre_val]
        total_conf = len(df_nre[df_nre['status'] != 'registered_only'])
        both_count = len(df_nre[df_nre['status'] == 'both'])
        faltam_count = len(df_nre[df_nre['status'] == 'confirmed_only'])
        pct_faltam = (faltam_count / total_conf * 100) if total_conf > 0 else 0
        nre_summary.append({
            'nre': nre_val,
            'total': total_conf,
            'ok': both_count,
            'faltam': faltam_count,
            'pct_faltam': pct_faltam
        })

    nre_summary.sort(key=lambda x: x['faltam'], reverse=True) # Mostrar os com mais faltantes no topo

    r_idx = start_row + 1
    for row_data in nre_summary:
        c1 = ws2.cell(row=r_idx, column=1, value=row_data['nre'])
        c2 = ws2.cell(row=r_idx, column=2, value=row_data['total'])
        c3 = ws2.cell(row=r_idx, column=3, value=row_data['ok'])
        c4 = ws2.cell(row=r_idx, column=4, value=row_data['faltam'])
        c5 = ws2.cell(row=r_idx, column=5, value=f"=D{r_idx}/B{r_idx}")

        c1.alignment = Alignment(horizontal="left")
        c2.alignment = Alignment(horizontal="right")
        c3.alignment = Alignment(horizontal="right")
        c4.alignment = Alignment(horizontal="right")
        c5.alignment = Alignment(horizontal="right")
        c5.number_format = '0.0%'

        for cell in [c1, c2, c3, c4, c5]:
            cell.font = regular_font
            cell.border = thin_border
        
        if row_data['faltam'] > 0:
            c4.font = bold_font
            c4.fill = red_badge_fill
            
        r_idx += 1

    # Linha Total NRE
    ws2.cell(row=r_idx, column=1, value="TOTAL GERAL").font = bold_font
    ws2.cell(row=r_idx, column=2, value=f"=SUM(B{start_row+1}:B{r_idx-1})").font = bold_font
    ws2.cell(row=r_idx, column=3, value=f"=SUM(C{start_row+1}:C{r_idx-1})").font = bold_font
    ws2.cell(row=r_idx, column=4, value=f"=SUM(D{start_row+1}:D{r_idx-1})").font = bold_font
    ws2.cell(row=r_idx, column=5, value=f"=D{r_idx}/B{r_idx}").font = bold_font
    ws2.cell(row=r_idx, column=5).number_format = '0.0%'

    for c in range(1, 6):
        cell = ws2.cell(row=r_idx, column=c)
        cell.fill = total_fill
        cell.border = thin_border

    for col in ws2.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws2.column_dimensions[col_letter].width = max(max_len + 4, 18)

    ws2.auto_filter.ref = f"A{start_row}:E{r_idx-1}"

    # ----------------------------------------------------
    # TAB 3: Visão Geral Completa
    # ----------------------------------------------------
    ws3 = wb.create_sheet(title="Todas Escolas Confirmadas")
    ws3.views.sheetView[0].showGridLines = True
    
    ws3.cell(row=1, column=1, value="Lista Completa de Escolas Confirmadas no Ofício").font = title_font
    ws3.cell(row=2, column=1, value="Status individual de cadastro no formulário").font = subtitle_font

    headers3 = ["Item", "NRE", "Município", "Nome da Escola", "Estudantes Previstos", "Status Cadastro"]
    
    start_row = 4
    for col_num, header_title in enumerate(headers3, 1):
        cell = ws3.cell(row=start_row, column=col_num, value=header_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[start_row].height = 28

    df_conf_all = df[df['status'] != 'registered_only'].sort_values(by=['nre', 'cidade_planilha', 'escola'])
    
    current_row = start_row + 1
    for idx, (_, r) in enumerate(df_conf_all.iterrows(), 1):
        c1 = ws3.cell(row=current_row, column=1, value=idx)
        c2 = ws3.cell(row=current_row, column=2, value=r['nre'])
        c3 = ws3.cell(row=current_row, column=3, value=r['cidade_planilha'])
        c4 = ws3.cell(row=current_row, column=4, value=r['escola'])
        c5 = ws3.cell(row=current_row, column=5, value=r['alunos'])
        
        is_ok = (r['status'] == 'both')
        c6 = ws3.cell(row=current_row, column=6, value="Cadastrada" if is_ok else "Pendente")

        c1.alignment = Alignment(horizontal="center")
        c2.alignment = Alignment(horizontal="left")
        c3.alignment = Alignment(horizontal="left")
        c4.alignment = Alignment(horizontal="left")
        c5.alignment = Alignment(horizontal="right")
        c6.alignment = Alignment(horizontal="center")

        if is_ok:
            c6.fill = green_badge_fill
            c6.font = green_badge_font
        else:
            c6.fill = red_badge_fill
            c6.font = red_badge_font

        for cell in [c1, c2, c3, c4, c5]:
            cell.font = regular_font
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = zebra_fill
        c6.border = thin_border
        
        ws3.row_dimensions[current_row].height = 20
        current_row += 1

    for col in ws3.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws3.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)
    ws3.column_dimensions['D'].width = 55

    ws3.auto_filter.ref = f"A{start_row}:F{current_row-1}"

    output_filename = "Escolas_Faltantes_Cadastro_TOEIC.xlsx"
    wb.save(output_filename)
    print(f"Arquivo gerado com sucesso: '{output_filename}'")

if __name__ == "__main__":
    generate_xlsx()
